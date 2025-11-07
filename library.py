#Create a repo for all the require library import
import pandas as pd
import numpy as np
import math 
import matplotlib.pyplot as plt
from datetime import datetime
import os
import shutil
import mysql.connector as sql
import time
from plotly.subplots import make_subplots
import plotly.graph_objects as go
import plotly.express as px
from sklearn.svm import LinearSVC
from sklearn.datasets import load_iris
from sklearn.datasets import make_classification
from sklearn.model_selection import train_test_split,cross_val_score
from sklearn.metrics import confusion_matrix,classification_report
from sklearn.preprocessing import MinMaxScaler,StandardScaler
from sklearn.decomposition import PCA
from sklearn.ensemble import RandomForestClassifier
from sklearn.datasets import make_classification
from imblearn.over_sampling import RandomOverSampler,SMOTE
from streamlit_extras.stylable_container import stylable_container
from jira import JIRA 
import streamlit as st
import configparser
from PIL import Image
from openpyxl import Workbook,load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment,PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from io import BytesIO
import io
import glob
from multiprocessing import Pool
import geopy.distance
from haversine import haversine, Unit
from copy import copy
from scipy.signal import savgol_filter
from sklearn.linear_model import RANSACRegressor
from scipy.ndimage import gaussian_filter1d
from typing import List, Union,Tuple
from scipy.ndimage import gaussian_filter1d
from concurrent.futures import ThreadPoolExecutor
import struct
import json
